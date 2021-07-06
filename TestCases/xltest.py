import openpyxl
from openpyxl.styles import PatternFill
import configparser
try:
    D1 = openpyxl.load_workbook("C:/Users/battu/Desktop/Data2.xlsx")
except FileNotFoundError:
     print("No File")
finally:
    print("just like that")
Sheets = D1.sheetnames
print(Sheets)
print(Sheets[0])
Sh1 = D1['Sheet1']
nor = Sh1.max_row
noc = Sh1.max_column
print (nor)
print (noc)
for i in range(1,nor+1):
    for j in range(1,noc+1):
        print(Sh1.cell(i,j).value)

Sh1.cell(row=5,column=1,value='suresh')
Sh1['A6']='Kumar'
Sh1['A6'].fill = PatternFill("solid",fgColor='71FF33')
D1.save("C:/Users/battu/Desktop/Data3.xlsx")